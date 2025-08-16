import openpyxl
import networkx as nx
import matplotlib.pyplot as plt
import re

def extrair_referencias(formula):
    """Extrai referências simples de célula da fórmula."""
    pattern = r'(?<![A-Z])(\$?[A-Z]{1,3}\$?\d+)'
    return re.findall(pattern, formula)

def construir_grafo_excel(arquivo_excel):
    wb = openpyxl.load_workbook(arquivo_excel, data_only=False)
    G = nx.DiGraph()

    for aba in wb.sheetnames:
        planilha = wb[aba]

        for linha in planilha.iter_rows():
            for celula in linha:
                if isinstance(celula.value, str) and celula.value.startswith("="):
                    origem = f"{aba}!{celula.coordinate}"
                    formula = celula.value

                    # Extrair referências
                    referencias = extrair_referencias(formula)

                    for ref in referencias:
                        destino = f"{aba}!{ref.replace('$', '')}"
                        G.add_node(destino)
                        G.add_node(origem)
                        G.add_edge(destino, origem)

                    # Marcar erros
                    if any(err in formula for err in ["#REF!", "#VALUE!", "#NAME?", "#DIV/0!"]):
                        G.nodes[origem]["erro"] = True
                    else:
                        G.nodes[origem]["erro"] = False

    return G

def plotar_grafo(G):
    plt.figure(figsize=(14, 10))
    pos = nx.spring_layout(G, k=0.6)

    # Separar nós com erro e sem erro
    erros = [n for n in G.nodes if G.nodes[n].get("erro")]
    corretos = [n for n in G.nodes if not G.nodes[n].get("erro")]

    nx.draw_networkx_nodes(G, pos, nodelist=corretos, node_color='skyblue', node_size=1000, label="Fórmulas corretas")
    nx.draw_networkx_nodes(G, pos, nodelist=erros, node_color='red', node_size=1000, label="Fórmulas com erro")

    nx.draw_networkx_edges(G, pos, arrowstyle='-|>', arrowsize=15)
    nx.draw_networkx_labels(G, pos, font_size=8)

    plt.title("Mapa de Dependências de Fórmulas no Excel", fontsize=14)
    plt.legend()
    plt.axis('off')
    plt.tight_layout()
    plt.show()

# === EXECUÇÃO ===
if __name__ == "__main__":
    caminho_arquivo = "C:\\Users\\ead\\Documents\\Linguagens de programação.xlsx"  # Altere para o nome do seu arquivo
    grafo = construir_grafo_excel(caminho_arquivo)
    print(f"Grafo com {grafo.number_of_nodes()} nós e {grafo.number_of_edges()} dependências.")
    plotar_grafo(grafo)
